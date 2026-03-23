from __future__ import annotations

from datetime import datetime, timedelta
from decimal import Decimal
from io import BytesIO

from aiogram import F, Router
from aiogram.filters import Command
from aiogram.types import BufferedInputFile, CallbackQuery, Message
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import and_, desc, select

from bot.keyboards import get_report_period_keyboard
from bot.services.finance import get_or_create_user
from config.i18n import get_text
from database.finance import convert_amount, normalize_currency
from database.group_context import get_active_group_id
from database.models import Category, Transaction, TransactionType, User
from database.session import async_session_factory

router = Router()

TYPE_LABELS = {
    'income': {'uz': 'Kirim', 'ru': 'Доход', 'en': 'Income'},
    'expense': {'uz': 'Chiqim', 'ru': 'Расход', 'en': 'Expense'},
    'transfer_in': {'uz': 'Kirim (transfer)', 'ru': 'Доход (перевод)', 'en': 'Income (transfer)'},
    'transfer_out': {'uz': 'Chiqim (transfer)', 'ru': 'Расход (перевод)', 'en': 'Expense (transfer)'},
}


async def _load_user(message: Message):
    src = message.from_user
    return await get_or_create_user(
        telegram_id=src.id,
        username=src.username,
        first_name=src.first_name,
        last_name=src.last_name,
        language_code=src.language_code,
    )


def _normalize_lang(value: str | None) -> str:
    lang = (value or 'uz').split('-')[0].lower()
    if lang in {'uz', 'ru', 'en'}:
        return lang
    return 'uz'


def _resolve_period_from_token(token: str | None) -> str | None:
    mapping = {
        'day': 'day',
        'daily': 'day',
        'week': 'week',
        'weekly': 'week',
        'month': 'month',
        'monthly': 'month',
        'year': 'year',
        'yearly': 'year',
    }
    if not token:
        return None
    return mapping.get(token.lower())


def _period_start(period: str) -> tuple[datetime, str]:
    now = datetime.now()
    if period == 'day':
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        return start, 'day'
    if period == 'week':
        start = (now - timedelta(days=now.weekday())).replace(
            hour=0, minute=0, second=0, microsecond=0
        )
        return start, 'week'
    if period == 'year':
        start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        return start, 'year'
    start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    return start, 'month'


def _localized_headers(lang: str) -> dict[str, str]:
    return {
        'title': {
            'uz': 'Moliyaviy hisobot',
            'ru': 'Финансовый отчёт',
            'en': 'Financial Report',
        }.get(lang, 'Financial Report'),
        'period': {'uz': 'Davr', 'ru': 'Период', 'en': 'Period'}.get(lang, 'Period'),
        'currency': {'uz': 'Valyuta', 'ru': 'Валюта', 'en': 'Currency'}.get(lang, 'Currency'),
        'generated': {
            'uz': 'Yaratilgan vaqt',
            'ru': 'Дата создания',
            'en': 'Generated at',
        }.get(lang, 'Generated at'),
        'income': {'uz': 'Jami kirim', 'ru': 'Итого доход', 'en': 'Total income'}.get(lang, 'Total income'),
        'expense': {'uz': 'Jami chiqim', 'ru': 'Итого расход', 'en': 'Total expense'}.get(lang, 'Total expense'),
        'difference': {'uz': 'Farq', 'ru': 'Разница', 'en': 'Difference'}.get(lang, 'Difference'),
        'details': {'uz': 'Barcha operatsiyalar', 'ru': 'Все операции', 'en': 'All operations'}.get(lang, 'All operations'),
        'col_date': {'uz': 'Sana', 'ru': 'Дата', 'en': 'Date'}.get(lang, 'Date'),
        'col_type': {'uz': 'Turi', 'ru': 'Тип', 'en': 'Type'}.get(lang, 'Type'),
        'col_category': {'uz': 'Kategoriya', 'ru': 'Категория', 'en': 'Category'}.get(lang, 'Category'),
        'col_amount': {'uz': 'Summa', 'ru': 'Сумма', 'en': 'Amount'}.get(lang, 'Amount'),
        'col_currency': {'uz': 'Valyuta', 'ru': 'Валюта', 'en': 'Currency'}.get(lang, 'Currency'),
        'col_amount_converted': {
            'uz': 'Summa (tanlangan valyuta)',
            'ru': 'Сумма (в выбранной валюте)',
            'en': 'Amount (selected currency)',
        }.get(lang, 'Amount (selected currency)'),
        'col_description': {'uz': 'Izoh', 'ru': 'Описание', 'en': 'Description'}.get(lang, 'Description'),
    }


def _build_workbook(
    *,
    lang: str,
    period_label: str,
    currency: str,
    total_income: Decimal,
    total_expense: Decimal,
    difference: Decimal,
    rows: list[dict],
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Summary'
    headers = _localized_headers(lang)

    title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    section_font = Font(name='Calibri', size=11, bold=True)
    header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    cell_font = Font(name='Calibri', size=10)

    title_fill = PatternFill('solid', fgColor='1F2937')
    header_fill = PatternFill('solid', fgColor='374151')
    alt_fill = PatternFill('solid', fgColor='F9FAFB')
    thin_side = Side(style='thin', color='D1D5DB')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    ws.merge_cells('A1:F1')
    ws['A1'] = headers['title']
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 28

    meta_rows = [
        (headers['period'], period_label),
        (headers['currency'], currency),
        (headers['generated'], datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        (headers['income'], float(total_income)),
        (headers['expense'], float(total_expense)),
        (headers['difference'], float(difference)),
    ]

    for row_idx, (label, value) in enumerate(meta_rows, start=3):
        ws.cell(row=row_idx, column=1, value=label).font = section_font
        value_cell = ws.cell(row=row_idx, column=2, value=value)
        value_cell.font = cell_font
        if isinstance(value, float):
            value_cell.number_format = '#,##0.00'

    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 24

    details = wb.create_sheet('Transactions')
    details_headers = [
        headers['col_date'],
        headers['col_type'],
        headers['col_category'],
        headers['col_amount'],
        headers['col_currency'],
        headers['col_amount_converted'],
        headers['col_description'],
    ]

    details.merge_cells('A1:G1')
    details['A1'] = headers['details']
    details['A1'].font = title_font
    details['A1'].fill = title_fill
    details['A1'].alignment = Alignment(horizontal='left', vertical='center')
    details.row_dimensions[1].height = 26

    for col_idx, title in enumerate(details_headers, start=1):
        cell = details.cell(row=3, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_idx, item in enumerate(rows, start=4):
        values = [
            item['date'],
            item['type_label'],
            item['category'],
            item['amount_original'],
            item['currency_original'],
            item['amount_converted'],
            item['description'],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = details.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.font = cell_font
            if col_idx in {4, 6}:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    details.auto_filter.ref = f"A3:G{max(3, len(rows) + 3)}"
    details.freeze_panes = 'A4'
    details.column_dimensions['A'].width = 20
    details.column_dimensions['B'].width = 20
    details.column_dimensions['C'].width = 24
    details.column_dimensions['D'].width = 14
    details.column_dimensions['E'].width = 12
    details.column_dimensions['F'].width = 20
    details.column_dimensions['G'].width = 48

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.read()


async def _generate_report_excel(user_id: int, period: str) -> tuple[bytes, str]:
    async with async_session_factory() as db:
        user = (await db.execute(select(User).where(User.id == user_id))).scalar_one_or_none()
        if not user:
            raise ValueError('User not found')

        now = datetime.now()
        start_date, label_type = _period_start(period)
        currency = normalize_currency(user.default_currency, 'UZS')
        lang = (user.language_code or 'uz').split('-')[0].lower()
        group_id = await get_active_group_id(db, user)

        if label_type == 'day':
            period_label = {'uz': 'Bugun', 'ru': 'Сегодня', 'en': 'Today'}.get(lang, 'Today')
        elif label_type == 'week':
            period_label = {'uz': 'Hafta', 'ru': 'Неделя', 'en': 'Week'}.get(lang, 'Week')
        elif label_type == 'year':
            period_label = str(start_date.year)
        else:
            period_label = now.strftime('%B %Y')

        tx_rows = (
            await db.execute(
                select(
                    Transaction.transaction_date,
                    Transaction.type,
                    Transaction.amount,
                    Transaction.currency,
                    Transaction.category_id,
                    Transaction.description,
                )
                .where(
                    and_(
                        Transaction.user_id == user.id,
                        Transaction.group_id == group_id,
                        Transaction.transaction_date >= start_date,
                    )
                )
                .order_by(desc(Transaction.transaction_date))
            )
        ).all()

        category_ids = [category_id for _, _, _, _, category_id, _ in tx_rows if category_id]
        category_map: dict[int, Category] = {}
        if category_ids:
            categories = (
                await db.execute(select(Category).where(Category.id.in_(list(set(category_ids)))))
            ).scalars().all()
            category_map = {cat.id: cat for cat in categories}

        total_income = Decimal('0')
        total_expense = Decimal('0')
        report_rows: list[dict] = []

        for tx_date, tx_type, amount, tx_currency, category_id, description in tx_rows:
            tx_type_value = tx_type.value if hasattr(tx_type, 'value') else str(tx_type)
            converted = await convert_amount(db, amount, tx_currency, currency)

            if tx_type_value in {TransactionType.INCOME.value, TransactionType.TRANSFER_IN.value}:
                total_income += converted
            elif tx_type_value in {TransactionType.EXPENSE.value, TransactionType.TRANSFER_OUT.value}:
                total_expense += converted

            category = category_map.get(category_id) if category_id else None
            category_title = f"{category.icon} {category.name}" if category else '-'
            type_label = TYPE_LABELS.get(tx_type_value, {}).get(lang, tx_type_value)

            report_rows.append(
                {
                    'date': tx_date.strftime('%Y-%m-%d %H:%M'),
                    'type_label': type_label,
                    'category': category_title,
                    'amount_original': float(amount),
                    'currency_original': tx_currency,
                    'amount_converted': float(converted),
                    'description': description or '-',
                }
            )

        difference = total_income - total_expense
        content = _build_workbook(
            lang=lang,
            period_label=period_label,
            currency=currency,
            total_income=total_income,
            total_expense=total_expense,
            difference=difference,
            rows=report_rows,
        )
        filename = f"report_{period}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return content, filename


async def _send_period_report(chat_message: Message, user_id: int, lang: str, period: str) -> None:
    progress = {
        'uz': '📄 Hisobot tayyorlanmoqda...',
        'ru': '📄 Подготавливаю отчёт...',
        'en': '📄 Preparing report...',
    }
    await chat_message.answer(progress.get(lang, progress['en']))

    try:
        file_bytes, filename = await _generate_report_excel(user_id, period)
    except Exception:
        error_text = {
            'uz': "❌ Hisobotni yaratib bo'lmadi",
            'ru': '❌ Не удалось сформировать отчёт',
            'en': '❌ Failed to generate report',
        }
        await chat_message.answer(error_text.get(lang, error_text['en']))
        return

    caption = {
        'uz': '✅ Excel hisobot tayyor',
        'ru': '✅ Excel-отчёт готов',
        'en': '✅ Excel report is ready',
    }

    await chat_message.answer_document(
        BufferedInputFile(file=file_bytes, filename=filename),
        caption=caption.get(lang, caption['en']),
    )


@router.message(Command('reports'))
@router.message(F.text.in_([get_text('btn_reports', 'uz'), get_text('btn_reports', 'ru'), get_text('btn_reports', 'en')]))
async def cmd_reports(message: Message):
    user = await _load_user(message)
    lang = _normalize_lang(user.language_code)

    parts = (message.text or '').strip().split()
    period = _resolve_period_from_token(parts[1] if len(parts) > 1 else None)

    if period:
        await _send_period_report(message, user.id, lang, period)
        return

    choose_text = {
        'uz': '📄 Hisobot davrini tanlang:',
        'ru': '📄 Выберите период отчёта:',
        'en': '📄 Choose report period:',
    }
    await message.answer(
        choose_text.get(lang, choose_text['en']),
        reply_markup=get_report_period_keyboard(lang),
    )


@router.callback_query(F.data.in_({'report_daily', 'report_weekly', 'report_monthly', 'report_yearly'}))
async def callback_reports_period(callback: CallbackQuery):
    user = await get_or_create_user(
        telegram_id=callback.from_user.id,
        username=callback.from_user.username,
        first_name=callback.from_user.first_name,
        last_name=callback.from_user.last_name,
        language_code=callback.from_user.language_code,
    )
    lang = _normalize_lang(user.language_code)

    period = _resolve_period_from_token(callback.data.replace('report_', '') if callback.data else None)
    if not period:
        await callback.answer('Invalid period', show_alert=True)
        return

    await callback.answer()
    await _send_period_report(callback.message, user.id, lang, period)
