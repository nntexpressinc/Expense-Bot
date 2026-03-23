from datetime import datetime, timedelta
from decimal import Decimal
from io import BytesIO
from typing import List

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from pydantic import BaseModel
from sqlalchemy import and_, desc, select
from sqlalchemy.ext.asyncio import AsyncSession

from api.routers.auth import get_current_user
from database.finance import convert_amount, normalize_currency
from database.group_context import get_active_group_id
from database.models import Category, Transaction, TransactionType, Debt, User
from database.session import get_db

router = APIRouter()


class CategoryStat(BaseModel):
    id: int
    name: str
    icon: str
    amount: float
    percent: float


class StatisticsResponse(BaseModel):
    period: str
    total_income: float
    total_expense: float
    difference: float
    top_categories: List[CategoryStat]


TYPE_LABELS = {
    'income': {'uz': 'Kirim', 'ru': 'Доход', 'en': 'Income'},
    'expense': {'uz': 'Chiqim', 'ru': 'Расход', 'en': 'Expense'},
    'transfer_in': {'uz': 'Kirim (transfer)', 'ru': 'Доход (перевод)', 'en': 'Income (transfer)'},
    'transfer_out': {'uz': 'Chiqim (transfer)', 'ru': 'Расход (перевод)', 'en': 'Expense (transfer)'},
    'debt': {'uz': 'Qarz olindi', 'ru': 'Долг взят', 'en': 'Debt taken'},
    'debt_payment': {'uz': 'Qarz to\'landi', 'ru': 'Погашение долга', 'en': 'Debt payment'},
}


def _period_start(period: str) -> tuple[datetime, str]:
    now = datetime.now()
    if period == 'day':
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        return start, 'day'
    if period == 'week':
        start = (now - timedelta(days=now.weekday())).replace(
            hour=0, minute=0, second=0, microsecond=0
        )
        return start, 'week'
    if period == 'year':
        start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
        return start, 'year'
    start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    return start, 'month'


def _period_label(lang: str, label_type: str, start: datetime, now: datetime) -> str:
    if label_type == 'day':
        return {
            'uz': f"Bugun ({start.strftime('%d.%m.%Y')})",
            'ru': f"Сегодня ({start.strftime('%d.%m.%Y')})",
            'en': f"Today ({start.strftime('%d.%m.%Y')})",
        }.get(lang, f"Bugun ({start.strftime('%d.%m.%Y')})")
    if label_type == 'week':
        return {
            'uz': f"Hafta ({start.strftime('%d.%m')} - {now.strftime('%d.%m.%Y')})",
            'ru': f"Неделя ({start.strftime('%d.%m')} - {now.strftime('%d.%m.%Y')})",
            'en': f"Week ({start.strftime('%d.%m')} - {now.strftime('%d.%m.%Y')})",
        }.get(lang, f"Hafta ({start.strftime('%d.%m')} - {now.strftime('%d.%m.%Y')})")
    if label_type == 'year':
        return {
            'uz': f"{start.year} yil",
            'ru': f"{start.year} год",
            'en': f"{start.year}",
        }.get(lang, f"{start.year} yil")
    month_name = start.strftime('%B %Y')
    return {'uz': month_name, 'ru': month_name, 'en': month_name}.get(lang, month_name)


def _localized_headers(lang: str) -> dict[str, str]:
    return {
        'title': {
            'uz': 'Moliyaviy hisobot',
            'ru': 'Финансовый отчёт',
            'en': 'Financial Report',
        }.get(lang, 'Financial Report'),
        'period': {'uz': 'Davr', 'ru': 'Период', 'en': 'Period'}.get(lang, 'Period'),
        'currency': {'uz': 'Valyuta', 'ru': 'Валюта', 'en': 'Currency'}.get(lang, 'Currency'),
        'generated': {
            'uz': 'Yaratilgan vaqt',
            'ru': 'Дата создания',
            'en': 'Generated at',
        }.get(lang, 'Generated at'),
        'income': {'uz': 'Jami kirim', 'ru': 'Итого доход', 'en': 'Total income'}.get(lang, 'Total income'),
        'expense': {'uz': 'Jami chiqim', 'ru': 'Итого расход', 'en': 'Total expense'}.get(lang, 'Total expense'),
        'difference': {'uz': 'Farq', 'ru': 'Разница', 'en': 'Difference'}.get(lang, 'Difference'),
        'details': {'uz': 'Barcha operatsiyalar', 'ru': 'Все операции', 'en': 'All operations'}.get(lang, 'All operations'),
        'col_date': {'uz': 'Sana', 'ru': 'Дата', 'en': 'Date'}.get(lang, 'Date'),
        'col_type': {'uz': 'Turi', 'ru': 'Тип', 'en': 'Type'}.get(lang, 'Type'),
        'col_category': {'uz': 'Kategoriya', 'ru': 'Категория', 'en': 'Category'}.get(lang, 'Category'),
        'col_amount': {'uz': 'Summa', 'ru': 'Сумма', 'en': 'Amount'}.get(lang, 'Amount'),
        'col_currency': {'uz': 'Valyuta', 'ru': 'Валюта', 'en': 'Currency'}.get(lang, 'Currency'),
        'col_amount_converted': {
            'uz': 'Summa (tanlangan valyuta)',
            'ru': 'Сумма (в выбранной валюте)',
            'en': 'Amount (selected currency)',
        }.get(lang, 'Amount (selected currency)'),
        'col_description': {'uz': 'Izoh', 'ru': 'Описание', 'en': 'Description'}.get(lang, 'Description'),
        'debts_sheet': {'uz': 'Qarzlar', 'ru': 'Долги', 'en': 'Debts'}.get(lang, 'Debts'),
        'debt_status': {'uz': 'Holat', 'ru': 'Статус', 'en': 'Status'}.get(lang, 'Status'),
        'debt_remaining': {'uz': 'Qolgan', 'ru': 'Остаток', 'en': 'Remaining'}.get(lang, 'Remaining'),
        'debt_created': {'uz': 'Yaratilgan', 'ru': 'Создан', 'en': 'Created'}.get(lang, 'Created'),
        'debt_paid': {'uz': "To'langan", 'ru': 'Погашен', 'en': 'Paid at'}.get(lang, 'Paid at'),
    }


def _build_statistics_workbook(
    *,
    lang: str,
    period_label: str,
    currency: str,
    total_income: Decimal,
    total_expense: Decimal,
    difference: Decimal,
    rows: list[dict],
    debt_rows: list[dict],
) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Summary'
    headers = _localized_headers(lang)

    title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    section_font = Font(name='Calibri', size=11, bold=True)
    header_font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
    cell_font = Font(name='Calibri', size=10)

    title_fill = PatternFill('solid', fgColor='1F2937')
    header_fill = PatternFill('solid', fgColor='374151')
    alt_fill = PatternFill('solid', fgColor='F9FAFB')
    thin_side = Side(style='thin', color='D1D5DB')
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    ws.merge_cells('A1:F1')
    ws['A1'] = headers['title']
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 28

    meta_rows = [
        (headers['period'], period_label),
        (headers['currency'], currency),
        (headers['generated'], datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        (headers['income'], float(total_income)),
        (headers['expense'], float(total_expense)),
        (headers['difference'], float(difference)),
    ]

    for row_idx, (label, value) in enumerate(meta_rows, start=3):
        ws.cell(row=row_idx, column=1, value=label).font = section_font
        value_cell = ws.cell(row=row_idx, column=2, value=value)
        value_cell.font = cell_font
        if isinstance(value, float):
            value_cell.number_format = '#,##0.00'

    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16

    details = wb.create_sheet('Transactions')
    details_headers = [
        headers['col_date'],
        headers['col_type'],
        headers['col_category'],
        headers['col_amount'],
        headers['col_currency'],
        headers['col_amount_converted'],
        headers['col_description'],
    ]

    details.merge_cells('A1:G1')
    details['A1'] = headers['details']
    details['A1'].font = title_font
    details['A1'].fill = title_fill
    details['A1'].alignment = Alignment(horizontal='left', vertical='center')
    details.row_dimensions[1].height = 26

    for col_idx, title in enumerate(details_headers, start=1):
        cell = details.cell(row=3, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_idx, item in enumerate(rows, start=4):
        values = [
            item['date'],
            item['type_label'],
            item['category'],
            item['amount_original'],
            item['currency_original'],
            item['amount_converted'],
            item['description'],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = details.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.font = cell_font
            if col_idx in {4, 6}:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    details.auto_filter.ref = f"A3:G{max(3, len(rows) + 3)}"
    details.freeze_panes = 'A4'
    details.column_dimensions['A'].width = 20
    details.column_dimensions['B'].width = 20
    details.column_dimensions['C'].width = 24
    details.column_dimensions['D'].width = 14
    details.column_dimensions['E'].width = 12
    details.column_dimensions['F'].width = 20
    details.column_dimensions['G'].width = 48

    # Debts sheet
    debts_ws = wb.create_sheet(headers['debts_sheet'])
    debts_headers = [
        headers['col_date'],
        headers['col_amount'],
        headers['col_currency'],
        headers['debt_remaining'],
        headers['debt_status'],
        headers['col_description'],
        headers['debt_paid'],
    ]
    debts_ws.merge_cells('A1:G1')
    debts_ws['A1'] = headers['debts_sheet']
    debts_ws['A1'].font = title_font
    debts_ws['A1'].fill = title_fill
    debts_ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    debts_ws.row_dimensions[1].height = 26

    for col_idx, title in enumerate(debts_headers, start=1):
        cell = debts_ws.cell(row=3, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_idx, item in enumerate(debt_rows, start=4):
        values = [
            item['created_at'],
            item['amount'],
            item['currency'],
            item['remaining'],
            item['status'],
            item['description'],
            item['paid_at'],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = debts_ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.font = cell_font
            if col_idx in {2, 4}:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    debts_ws.auto_filter.ref = f"A3:G{max(3, len(debt_rows) + 3)}"
    debts_ws.freeze_panes = 'A4'
    for key, width in zip("ABCDEFG", [20, 14, 10, 14, 14, 40, 22]):
        debts_ws.column_dimensions[key].width = width

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


@router.get('/export/excel')
async def export_statistics_excel(
    period: str = Query('month', pattern='^(day|week|month|year)$'),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    now = datetime.now()
    start_date, label_type = _period_start(period)
    currency = normalize_currency(current_user.default_currency, 'UZS')
    lang = (current_user.language_code or 'uz').split('-')[0].lower()
    period_label = _period_label(lang, label_type, start_date, now)
    group_id = await get_active_group_id(db, current_user)

    tx_rows = (
        await db.execute(
            select(
                Transaction.transaction_date,
                Transaction.type,
                Transaction.amount,
                Transaction.currency,
                Transaction.category_id,
                Transaction.description,
            )
            .where(
                and_(
                    Transaction.user_id == current_user.id,
                    Transaction.group_id == group_id,
                    Transaction.transaction_date >= start_date,
                )
            )
            .order_by(desc(Transaction.transaction_date))
        )
    ).all()

    debt_rows_raw = (
        await db.execute(
            select(Debt)
            .where(Debt.user_id == current_user.id, Debt.group_id == group_id)
            .order_by(desc(Debt.created_at))
        )
    ).scalars().all()

    category_ids = [category_id for _, _, _, _, category_id, _ in tx_rows if category_id]
    category_map: dict[int, Category] = {}
    if category_ids:
        categories = (
            await db.execute(select(Category).where(Category.id.in_(list(set(category_ids)))))
        ).scalars().all()
        category_map = {cat.id: cat for cat in categories}

    total_income = Decimal('0')
    total_expense = Decimal('0')
    report_rows: list[dict] = []

    for tx_date, tx_type, amount, tx_currency, category_id, description in tx_rows:
        tx_type_value = tx_type.value if hasattr(tx_type, 'value') else str(tx_type)
        converted = await convert_amount(db, amount, tx_currency, currency)

        if tx_type_value in {TransactionType.INCOME.value, TransactionType.TRANSFER_IN.value}:
            total_income += converted
        elif tx_type_value in {
            TransactionType.EXPENSE.value,
            TransactionType.TRANSFER_OUT.value,
            TransactionType.DEBT_PAYMENT.value,
        }:
            total_expense += converted

        category = category_map.get(category_id) if category_id else None
        category_title = f"{category.icon} {category.name}" if category else '-'
        type_label = TYPE_LABELS.get(tx_type_value, {}).get(lang, tx_type_value)

        report_rows.append(
            {
                'date': tx_date.strftime('%Y-%m-%d %H:%M'),
                'type_label': type_label,
                'category': category_title,
                'amount_original': float(amount),
                'currency_original': tx_currency,
                'amount_converted': float(converted),
                'description': description or '-',
            }
        )

    difference = total_income - total_expense
    debt_rows_payload = []
    for debt in debt_rows_raw:
        debt_rows_payload.append(
            {
                'created_at': debt.created_at.strftime('%Y-%m-%d %H:%M'),
                'amount': float(debt.amount),
                'currency': debt.currency,
                'remaining': float(debt.remaining_amount),
                'status': 'paid' if debt.remaining_amount == 0 else 'open',
                'description': debt.description or '-',
                'paid_at': debt.paid_at.strftime('%Y-%m-%d %H:%M') if debt.paid_at else '-',
            }
        )

    file_stream = _build_statistics_workbook(
        lang=lang,
        period_label=period_label,
        currency=currency,
        total_income=total_income,
        total_expense=total_expense,
        difference=difference,
        rows=report_rows,
        debt_rows=debt_rows_payload,
    )
    filename = f"statistics_{period}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        file_stream,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


@router.get('/', response_model=StatisticsResponse)
async def get_statistics(
    period: str = Query('month', pattern='^(day|week|month|year)$'),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    now = datetime.now()
    start_date, label_type = _period_start(period)
    currency = normalize_currency(current_user.default_currency, 'UZS')
    lang = (current_user.language_code or 'uz').split('-')[0].lower()
    group_id = await get_active_group_id(db, current_user)

    rows = (
        await db.execute(
            select(Transaction.type, Transaction.amount, Transaction.currency, Transaction.category_id)
            .where(
                and_(
                    Transaction.user_id == current_user.id,
                    Transaction.group_id == group_id,
                    Transaction.transaction_date >= start_date,
                )
            )
        )
    ).all()

    total_income = Decimal('0')
    total_expense = Decimal('0')
    category_totals: dict[int, Decimal] = {}

    for tx_type, amount, tx_currency, category_id in rows:
        tx_type_val = tx_type.value if hasattr(tx_type, 'value') else str(tx_type)
        converted = await convert_amount(db, amount, tx_currency, currency)

        if tx_type_val == TransactionType.INCOME.value:
            total_income += converted
        elif tx_type_val in {
            TransactionType.EXPENSE.value,
            TransactionType.TRANSFER_OUT.value,
            TransactionType.DEBT_PAYMENT.value,
        }:
            total_expense += converted
            if tx_type_val == TransactionType.EXPENSE.value and category_id:
                category_totals[category_id] = category_totals.get(category_id, Decimal('0')) + converted

    top_categories: List[dict] = []
    if category_totals:
        category_ids = list(category_totals.keys())
        categories = (
            await db.execute(select(Category).where(Category.id.in_(category_ids)))
        ).scalars().all()
        category_map = {cat.id: cat for cat in categories}

        ordered = sorted(category_totals.items(), key=lambda item: item[1], reverse=True)[:5]
        for cat_id, amount_value in ordered:
            cat = category_map.get(cat_id)
            percent = float((amount_value / total_expense * 100) if total_expense > 0 else Decimal('0'))
            top_categories.append(
                {
                    'id': cat_id,
                    'name': cat.name if cat else 'Unknown',
                    'icon': cat.icon if cat else '💰',
                    'amount': float(amount_value),
                    'percent': round(percent, 1),
                }
            )

    return {
        'period': _period_label(lang, label_type, start_date, now),
        'total_income': float(total_income),
        'total_expense': float(total_expense),
        'difference': float(total_income - total_expense),
        'top_categories': top_categories,
    }
