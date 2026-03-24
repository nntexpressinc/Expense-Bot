from io import BytesIO
import zipfile
from types import SimpleNamespace

from openpyxl import load_workbook
import pytest

from database.reporting import build_excel_workbook, generate_report_download


def _sample_payload(admin_mode: bool):
    return {
        "group_name": "Toshkent Office",
        "period_label": "March 2026",
        "currency": "USD",
        "summary": [
            ["Group", "Toshkent Office"],
            ["User", "Xolmirza"],
            ["Total income", 120.0],
            ["Total expense", 45.0],
        ],
        "top_categories": [["Food", 30.0, 66.7]],
        "transactions": [
            {
                "date": "2026-03-24 10:00",
                "type": "Expense",
                "category": "Food",
                "funding": "Debt source",
                "source": "Xolmirza",
                "amount": 30.0,
                "currency": "USD",
                "converted": 30.0,
                "description": "Lunch",
            }
        ],
        "debts": [["2026-03-24 09:00", "Borrow money", "Xolmirza", "-", 100.0, "USD", 70.0, 30.0, 70.0, "active", "Borrowed", "-"]],
        "debt_repayments": [["2026-03-24 12:00", "Borrowed", "Xolmirza", 10.0, "USD", 10.0, "-"]],
        "admin_mode": admin_mode,
        "users": [["Xolmirza", "@xolmirza", "admin", "Yes", "uz", "USD", "2026-03-24 09:00"]],
        "transfers": [["2026-03-24 08:00", "Admin", "User", 50.0, "USD", 20.0, 30.0, "completed", "Office money"]],
        "workers": [["Ali", "Farmer", "daily", 10.0, "USD", "Yes", "2026-03-01", 20.0, 200.0, 50.0, 100.0, 50.0, "partial", "-", "-"]],
        "attendance": [["2026-03-24", "Ali", "daily", "present", 1.0, "-"]],
        "advances": [["2026-03-10", "Ali", 50.0, "USD", 50.0, "-"]],
        "worker_payments": [["2026-03-20", "Ali", 100.0, "USD", 100.0, "-"]],
        "audit": [["2026-03-24 12:10", "Xolmirza", "debt.created", "debt", "1", "{}"]],
    }


def test_build_excel_workbook_basic_sheets():
    workbook = load_workbook(BytesIO(build_excel_workbook(_sample_payload(admin_mode=False))))

    assert workbook.sheetnames == ["Summary", "Transactions", "Debts", "Debt Repayments"]
    assert workbook["Summary"]["A1"].value == "Financial Monitoring Report"
    assert workbook["Transactions"]["A4"].value == "2026-03-24 10:00"
    assert workbook["Transactions"]["A5"].value == "Total"
    assert workbook["Transactions"]["F5"].value == 30
    assert workbook["Debt Repayments"]["A5"].value == "Total"


def test_build_excel_workbook_admin_monitoring_sheets():
    workbook = load_workbook(BytesIO(build_excel_workbook(_sample_payload(admin_mode=True))))

    assert "Users" in workbook.sheetnames
    assert "Transfers" in workbook.sheetnames
    assert "Workers" in workbook.sheetnames
    assert "Attendance" in workbook.sheetnames
    assert "Advances" in workbook.sheetnames
    assert "Worker Payments" in workbook.sheetnames
    assert "Audit Log" in workbook.sheetnames
    assert workbook["Workers"]["A4"].value == "Ali"
    assert workbook["Workers"]["A5"].value == "Total"


@pytest.mark.asyncio
async def test_generate_report_download_admin_bundle_contains_group_and_member_files(monkeypatch):
    async def fake_is_group_admin(*_args, **_kwargs):
        return True

    async def fake_get_active_group_id(*_args, **_kwargs):
        return 7

    async def fake_get_active_group(*_args, **_kwargs):
        return SimpleNamespace(name="Toshkent Office")

    async def fake_generate_excel_report(_db, user, _period, *, filename_prefix="report", **_kwargs):
        safe_name = str(getattr(user, "id", "group"))
        return b"xlsx", f"{filename_prefix}_{safe_name}.xlsx"

    class _ScalarResult:
        def __init__(self, rows):
            self._rows = rows

        def all(self):
            return self._rows

    class _ExecuteResult:
        def __init__(self, rows):
            self._rows = rows

        def scalars(self):
            return _ScalarResult(self._rows)

    class DummyDb:
        async def execute(self, _query):
            members = [
                SimpleNamespace(id=10, first_name="Ali", last_name=None, username="ali"),
                SimpleNamespace(id=11, first_name="Vali", last_name=None, username="vali"),
            ]
            return _ExecuteResult(members)

    monkeypatch.setattr("database.reporting.is_group_admin", fake_is_group_admin)
    monkeypatch.setattr("database.reporting.get_active_group_id", fake_get_active_group_id)
    monkeypatch.setattr("database.reporting.get_active_group", fake_get_active_group)
    monkeypatch.setattr("database.reporting.generate_excel_report", fake_generate_excel_report)

    content, filename, media_type = await generate_report_download(
        DummyDb(),
        SimpleNamespace(id=1, first_name="Admin", last_name=None, username="admin"),
        "month",
        filename_prefix="statistics",
    )

    archive = zipfile.ZipFile(BytesIO(content))
    names = set(archive.namelist())

    assert media_type == "application/zip"
    assert filename.endswith(".zip")
    assert "README.txt" in names
    assert any(name.startswith("statistics_group_") for name in names)
    assert any(name.startswith("statistics_ali_10_") for name in names)
    assert any(name.startswith("statistics_vali_11_") for name in names)
