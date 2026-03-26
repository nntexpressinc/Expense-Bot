from __future__ import annotations

import re
import zipfile
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import and_, desc, select
from sqlalchemy.ext.asyncio import AsyncSession

from database.finance import convert_amount, get_available_debt_for_entry, get_user_balance_summary, normalize_currency
from database.group_context import get_active_group, get_active_group_id, is_group_admin
from database.models import (
    AttendanceEntry,
    AuditLog,
    Category,
    Debt,
    DebtRepayment,
    DebtUsage,
    Group,
    Transaction,
    TransactionType,
    Transfer,
    User,
    UserGroup,
    Worker,
    WorkerAdvance,
    WorkerPayment,
)
from database.workers import calculate_group_payroll_summary


def _period_start(period: str) -> tuple[datetime, str]:
    now = datetime.now()
    if period == "day":
        return now.replace(hour=0, minute=0, second=0, microsecond=0), "day"
    if period == "week":
        return (now - timedelta(days=now.weekday())).replace(hour=0, minute=0, second=0, microsecond=0), "week"
    if period == "year":
        return now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0), "year"
    return now.replace(day=1, hour=0, minute=0, second=0, microsecond=0), "month"


def _period_bounds(
    period: str,
    *,
    date_from: date | None = None,
    date_to: date | None = None,
) -> tuple[datetime, datetime, str]:
    now = datetime.now()
    if period == "custom":
        if date_from is None or date_to is None:
            raise ValueError("Custom period requires date_from and date_to")
        start = datetime.combine(date_from, time.min)
        end = datetime.combine(date_to, time.max)
        if start > end:
            raise ValueError("date_from must be before or equal to date_to")
        return start, end, "custom"

    start, period_type = _period_start(period)
    return start, now, period_type


def _period_label(period_type: str, start: datetime, now: datetime, end: datetime | None = None) -> str:
    if period_type == "day":
        return f"Today ({start:%d.%m.%Y})"
    if period_type == "week":
        return f"Week ({start:%d.%m} - {now:%d.%m.%Y})"
    if period_type == "year":
        return f"{start.year}"
    if period_type == "custom":
        range_end = end or now
        return f"{start:%d.%m.%Y} - {range_end:%d.%m.%Y}"
    return start.strftime("%B %Y")


def _display_name(user: User) -> str:
    full_name = f"{user.first_name or ''} {user.last_name or ''}".strip()
    if full_name:
        return full_name
    if user.username:
        return f"@{user.username}"
    return str(user.id)


def _safe_filename_part(value: str) -> str:
    cleaned = re.sub(r"[^a-zA-Z0-9_-]+", "-", (value or "").strip().lower()).strip("-")
    return cleaned or "group"


def _is_synthetic_opening_debt(debt: Debt) -> bool:
    return (debt.reference or "") == "OPENING-DEBT-RANCHO" or (debt.description or "").startswith(
        "Opening debt carry-over"
    )


def _type_label(tx_type_value: str, debt_kind: str | None = None) -> str:
    if tx_type_value == TransactionType.DEBT.value:
        return "Money borrowed" if debt_kind == "cash_loan" else "Buy on credit"
    mapping = {
        TransactionType.INCOME.value: "Income",
        TransactionType.EXPENSE.value: "Expense",
        TransactionType.TRANSFER_IN.value: "Transfer in",
        TransactionType.TRANSFER_OUT.value: "Transfer out",
        TransactionType.DEBT_PAYMENT.value: "Debt payment",
    }
    return mapping.get(tx_type_value, tx_type_value)


def _styles() -> dict[str, Any]:
    thin = Side(style="thin", color="D1D5DB")
    return {
        "title_font": Font(name="Calibri", size=15, bold=True, color="FFFFFF"),
        "section_font": Font(name="Calibri", size=11, bold=True, color="111827"),
        "header_font": Font(name="Calibri", size=10, bold=True, color="FFFFFF"),
        "cell_font": Font(name="Calibri", size=10, color="111827"),
        "title_fill": PatternFill("solid", fgColor="111827"),
        "header_fill": PatternFill("solid", fgColor="374151"),
        "muted_fill": PatternFill("solid", fgColor="F3F4F6"),
        "alt_fill": PatternFill("solid", fgColor="FAFAFA"),
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
    }


def _write_sheet_title(ws, title: str, last_col: int, styles: dict[str, Any]) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = styles["title_font"]
    cell.fill = styles["title_fill"]
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26


def _write_table(
    ws,
    *,
    start_row: int,
    headers: list[str],
    rows: list[list[Any]],
    widths: list[int],
    numeric_cols: set[int] | None = None,
) -> None:
    styles = _styles()
    numeric_cols = numeric_cols or set()
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=title)
        cell.font = styles["header_font"]
        cell.fill = styles["header_fill"]
        cell.border = styles["border"]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_idx, row in enumerate(rows, start=start_row + 1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = styles["cell_font"]
            cell.border = styles["border"]
            if col_idx in numeric_cols and isinstance(value, (int, float, Decimal)):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if row_idx % 2 == 0:
                cell.fill = styles["alt_fill"]
    end_col = chr(64 + len(headers))
    ws.auto_filter.ref = f"A{start_row}:{end_col}{max(start_row, start_row + len(rows))}"
    ws.freeze_panes = f"A{start_row + 1}"
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width


def _append_total_row(rows: list[list[Any]], label: str, numeric_cols: set[int]) -> list[list[Any]]:
    if not rows:
        return rows

    totals: list[Any] = ["" for _ in rows[0]]
    totals[0] = label
    has_values = False
    for col_idx in numeric_cols:
        row_index = col_idx - 1
        total_value = Decimal("0")
        for row in rows:
            if row_index >= len(row):
                continue
            value = row[row_index]
            if isinstance(value, (int, float, Decimal)):
                total_value += Decimal(str(value))
        if total_value:
            totals[row_index] = float(total_value)
            has_values = True
    return rows + [totals] if has_values else rows


async def collect_excel_report_payload(
    db: AsyncSession,
    user: User,
    period: str,
    *,
    date_from: date | None = None,
    date_to: date | None = None,
) -> dict[str, Any]:
    now = datetime.now()
    start_date, end_date, period_type = _period_bounds(period, date_from=date_from, date_to=date_to)
    currency = normalize_currency(user.default_currency, "UZS")
    group_id = await get_active_group_id(db, user)
    group = await get_active_group(db, user)
    admin_mode = await is_group_admin(db, user, group_id)
    return await collect_excel_report_payload_for_group(
        db,
        user,
        period,
        group_id=group_id,
        include_admin_sheets=admin_mode,
        group=group,
        now=now,
        start_date=start_date,
        end_date=end_date,
        period_type=period_type,
        currency=currency,
    )


async def collect_excel_report_payload_for_group(
    db: AsyncSession,
    user: User,
    period: str,
    *,
    group_id: int,
    include_admin_sheets: bool,
    group=None,
    now: datetime | None = None,
    start_date: datetime | None = None,
    end_date: datetime | None = None,
    period_type: str | None = None,
    currency: str | None = None,
) -> dict[str, Any]:
    now = now or datetime.now()
    if start_date and end_date and period_type:
        pass
    else:
        start_date, end_date, period_type = _period_bounds(period)
    currency = currency or normalize_currency(user.default_currency, "UZS")
    if group is None:
        group = (await db.execute(select(Group).where(Group.id == group_id))).scalar_one()
    admin_mode = include_admin_sheets
    tx_rows = (
        await db.execute(
            select(
                Transaction.id,
                Transaction.transaction_date,
                Transaction.type,
                Transaction.debt_kind,
                Transaction.amount,
                Transaction.currency,
                Transaction.category_id,
                Transaction.description,
                Transaction.funding_source,
            )
            .where(
                and_(
                    Transaction.user_id == user.id,
                    Transaction.group_id == group_id,
                    Transaction.transaction_date >= start_date,
                    Transaction.transaction_date <= end_date,
                )
            )
            .order_by(desc(Transaction.transaction_date))
        )
    ).all()
    tx_ids = [row[0] for row in tx_rows]
    usage_map: dict[Any, str] = {}
    if tx_ids:
        usage_rows = (
            await db.execute(
                select(DebtUsage.transaction_id, Debt.source_name, Debt.description)
                .join(Debt, Debt.id == DebtUsage.debt_id)
                .where(DebtUsage.transaction_id.in_(tx_ids))
            )
        ).all()
        usage_map = {tx_id: (source_name or description or "-") for tx_id, source_name, description in usage_rows}

    category_ids = sorted({row[6] for row in tx_rows if row[6]})
    category_map: dict[int, Category] = {}
    if category_ids:
        categories = (await db.execute(select(Category).where(Category.id.in_(category_ids)))).scalars().all()
        category_map = {category.id: category for category in categories}

    total_income = Decimal("0")
    total_expense = Decimal("0")
    category_totals: dict[int, Decimal] = {}
    income_category_totals: dict[str, Decimal] = {}
    transaction_rows: list[dict[str, Any]] = []
    for tx_id, tx_date, tx_type, debt_kind, amount, tx_currency, category_id, description, funding_source in tx_rows:
        tx_type_value = tx_type.value if hasattr(tx_type, "value") else str(tx_type)
        category = category_map.get(category_id) if category_id else None
        converted = await convert_amount(db, amount, tx_currency, currency)
        if tx_type_value in {TransactionType.INCOME.value, TransactionType.TRANSFER_IN.value}:
            total_income += converted
            income_label = "Transfers received" if tx_type_value == TransactionType.TRANSFER_IN.value else (
                category.name if category else _type_label(tx_type_value, debt_kind)
            )
            income_category_totals[income_label] = income_category_totals.get(income_label, Decimal("0")) + converted
        elif tx_type_value in {TransactionType.EXPENSE.value, TransactionType.TRANSFER_OUT.value, TransactionType.DEBT_PAYMENT.value}:
            total_expense += converted
            if tx_type_value == TransactionType.EXPENSE.value and category_id:
                category_totals[category_id] = category_totals.get(category_id, Decimal("0")) + converted

        transaction_rows.append(
            {
                "date": tx_date.strftime("%Y-%m-%d %H:%M"),
                "type": _type_label(tx_type_value, debt_kind),
                "category": f"{category.icon} {category.name}" if category else "-",
                "funding": "Debt source" if funding_source == "debt" else "Main balance",
                "source": usage_map.get(tx_id, "-") if funding_source == "debt" else "-",
                "amount": float(amount),
                "currency": tx_currency,
                "converted": float(converted),
                "description": description or "-",
            }
        )

    top_categories = []
    if category_totals:
        ordered = sorted(category_totals.items(), key=lambda item: item[1], reverse=True)[:5]
        for category_id, amount_value in ordered:
            category = category_map.get(category_id)
            percent = float((amount_value / total_expense * 100) if total_expense > 0 else Decimal("0"))
            top_categories.append([f"{category.icon} {category.name}" if category else "-", float(amount_value), round(percent, 1)])

    income_breakdown_rows = []
    if income_category_totals:
        ordered_income = sorted(income_category_totals.items(), key=lambda item: item[1], reverse=True)
        for label, amount_value in ordered_income[:5]:
            income_breakdown_rows.append([f"Income: {label}", float(amount_value)])

    debts = (
        await db.execute(
            select(Debt)
            .where(
                Debt.user_id == user.id,
                Debt.group_id == group_id,
                Debt.created_at <= end_date,
            )
            .order_by(desc(Debt.created_at))
        )
    ).scalars().all()
    reportable_debts = [debt for debt in debts if not _is_synthetic_opening_debt(debt)]
    debt_total = Decimal("0")
    debt_rows = []
    for debt in reportable_debts:
        converted_amount = await convert_amount(db, debt.amount, debt.currency, currency)
        debt_total += converted_amount
        debt_rows.append(
            [
                debt.created_at.strftime("%Y-%m-%d %H:%M"),
                "Borrow money" if debt.kind == "cash_loan" else "Buy on credit",
                debt.source_name or "-",
                debt.source_contact or "-",
                float(converted_amount),
                currency,
                float(await convert_amount(db, debt.remaining_amount, debt.currency, currency)),
                float(await convert_amount(db, debt.used_amount, debt.currency, currency)),
                float(await get_available_debt_for_entry(db, debt, currency)),
                debt.status,
                debt.description or "-",
                debt.paid_at.strftime("%Y-%m-%d %H:%M") if debt.paid_at else "-",
            ]
        )

    debt_ids = [debt.id for debt in debts]
    repayment_rows = []
    if debt_ids:
        repayments = (
            await db.execute(
                select(DebtRepayment, Debt)
                .join(Debt, Debt.id == DebtRepayment.debt_id)
                .where(
                    DebtRepayment.debt_id.in_(debt_ids),
                    DebtRepayment.repaid_at >= start_date,
                    DebtRepayment.repaid_at <= end_date,
                )
                .order_by(desc(DebtRepayment.repaid_at))
            )
        ).all()
        for repayment, debt in repayments:
            note = repayment.note or "-"
            if note and "record=" in note and "source_row=" in note:
                note = debt.description or note.split("|", 1)[0].strip()
            repayment_rows.append(
                [
                    repayment.repaid_at.strftime("%Y-%m-%d %H:%M"),
                    ("Historical debt carry-over" if _is_synthetic_opening_debt(debt) else (debt.description or str(debt.id))),
                    debt.source_name or "-",
                    float(repayment.amount),
                    repayment.currency,
                    float(await convert_amount(db, repayment.amount, repayment.currency, currency)),
                    note,
                ]
            )

    balance_summary = await get_user_balance_summary(db, user, currency, group_id)
    summary_rows = [
        ["Group", group.name],
        ["User", _display_name(user)],
        ["Role", "Super admin" if user.is_admin else ("Group admin" if admin_mode else "User")],
        ["Period", _period_label(period_type, start_date, now, end_date)],
        ["Currency", currency],
        ["Generated at", now.strftime("%Y-%m-%d %H:%M:%S")],
        ["Main balance", float(balance_summary["total_balance"])],
        ["Own balance", float(balance_summary["own_balance"])],
        ["Transfer balance", float(balance_summary["received_balance"])],
    ]
    summary_rows.extend(income_breakdown_rows)
    summary_rows.extend([
        ["All income", float(total_income)],
        ["Total expense", float(total_expense)],
        ["Debt taken", float(debt_total)],
        ["Net result", float(total_income - total_expense)],
        ["Debt balance", float(balance_summary["debt_balance"])],
        ["Outstanding debt", float(balance_summary["outstanding_debt_balance"])],
        ["Transactions count", len(transaction_rows)],
        ["Debts count", len(reportable_debts)],
        ["Debt repayments", len(repayment_rows)],
    ])

    admin_payload: dict[str, list[list[Any]]] = {
        "users": [],
        "transfers": [],
        "workers": [],
        "attendance": [],
        "advances": [],
        "worker_payments": [],
        "audit": [],
    }
    if admin_mode:
        members = (
            await db.execute(
                select(UserGroup, User).join(User, User.id == UserGroup.user_id).where(UserGroup.group_id == group_id)
            )
        ).all()
        admin_payload["users"] = [
            [
                _display_name(member),
                f"@{member.username}" if member.username else "-",
                membership.role,
                "Yes" if member.is_active else "No",
                member.language_code or "-",
                member.default_currency,
                membership.joined_at.strftime("%Y-%m-%d %H:%M") if membership.joined_at else "-",
            ]
            for membership, member in members
        ]

        transfers = (
            await db.execute(
                select(Transfer)
                .where(
                    Transfer.group_id == group_id,
                    Transfer.created_at >= start_date,
                    Transfer.created_at <= end_date,
                )
                .order_by(desc(Transfer.created_at))
            )
        ).scalars().all()
        user_ids = {transfer.sender_id for transfer in transfers} | {transfer.recipient_id for transfer in transfers}
        user_map = {}
        if user_ids:
            users = (await db.execute(select(User).where(User.id.in_(list(user_ids))))).scalars().all()
            user_map = {member.id: _display_name(member) for member in users}
        admin_payload["transfers"] = [
            [
                transfer.created_at.strftime("%Y-%m-%d %H:%M"),
                user_map.get(transfer.sender_id, str(transfer.sender_id)),
                user_map.get(transfer.recipient_id, str(transfer.recipient_id)),
                float(await convert_amount(db, transfer.amount, transfer.currency, currency)),
                currency,
                float(await convert_amount(db, transfer.remaining_amount, transfer.currency, currency)),
                float(await convert_amount(db, transfer.amount - transfer.remaining_amount, transfer.currency, currency)),
                transfer.status.value if hasattr(transfer.status, "value") else str(transfer.status),
                transfer.description or "-",
            ]
            for transfer in transfers
        ]

        payroll = await calculate_group_payroll_summary(
            db,
            group_id=group_id,
            start_date=start_date.date(),
            end_date=end_date.date(),
            target_currency=currency,
            include_inactive=True,
        )
        summary_rows.extend(
            [
                ["Transfers count", len(admin_payload["transfers"])],
                ["Workers count", len(payroll["workers"])],
                ["Payroll accrued", payroll["totals"]["base_amount"]],
                ["Payroll advances", payroll["totals"]["advance_amount"]],
                ["Payroll paid", payroll["totals"]["paid_amount"]],
                ["Payroll payable", payroll["totals"]["payable_amount"]],
            ]
        )

        worker_map = {item["worker_id"]: item for item in payroll["workers"]}
        workers = (await db.execute(select(Worker).where(Worker.group_id == group_id).order_by(Worker.full_name.asc()))).scalars().all()
        admin_payload["workers"] = [
            [
                worker.full_name,
                worker.role_name or "-",
                worker.payment_type,
                worker_map.get(str(worker.id), {}).get("rate", float(worker.rate)),
                currency,
                "Yes" if worker.is_active else "No",
                worker.start_date.isoformat(),
                worker_map.get(str(worker.id), {}).get("quantity", 0),
                worker_map.get(str(worker.id), {}).get("base_amount", 0),
                worker_map.get(str(worker.id), {}).get("advance_amount", 0),
                worker_map.get(str(worker.id), {}).get("paid_amount", 0),
                worker_map.get(str(worker.id), {}).get("payable_amount", 0),
                worker_map.get(str(worker.id), {}).get("status", "-"),
                worker.phone or "-",
                worker.notes or "-",
            ]
            for worker in workers
        ]

        attendance = (
            await db.execute(
                select(AttendanceEntry, Worker)
                .join(Worker, Worker.id == AttendanceEntry.worker_id)
                .where(
                    AttendanceEntry.group_id == group_id,
                    AttendanceEntry.entry_date >= start_date.date(),
                    AttendanceEntry.entry_date <= end_date.date(),
                )
                .order_by(desc(AttendanceEntry.entry_date), Worker.full_name.asc())
            )
        ).all()
        admin_payload["attendance"] = [
            [entry.entry_date.isoformat(), worker.full_name, worker.payment_type, entry.status, float(entry.units), entry.comment or "-"]
            for entry, worker in attendance
        ]

        advances = (
            await db.execute(
                select(WorkerAdvance, Worker)
                .join(Worker, Worker.id == WorkerAdvance.worker_id)
                .where(
                    WorkerAdvance.group_id == group_id,
                    WorkerAdvance.payment_date >= start_date.date(),
                    WorkerAdvance.payment_date <= end_date.date(),
                )
                .order_by(desc(WorkerAdvance.payment_date))
            )
        ).all()
        admin_payload["advances"] = [
            [
                advance.payment_date.isoformat(),
                worker.full_name,
                float(advance.amount),
                advance.currency,
                float(await convert_amount(db, advance.amount, advance.currency, currency)),
                advance.note or "-",
            ]
            for advance, worker in advances
        ]

        payments = (
            await db.execute(
                select(WorkerPayment, Worker)
                .join(Worker, Worker.id == WorkerPayment.worker_id)
                .where(
                    WorkerPayment.group_id == group_id,
                    WorkerPayment.payment_date >= start_date.date(),
                    WorkerPayment.payment_date <= end_date.date(),
                )
                .order_by(desc(WorkerPayment.payment_date))
            )
        ).all()
        admin_payload["worker_payments"] = [
            [
                payment.payment_date.isoformat(),
                worker.full_name,
                float(payment.amount),
                payment.currency,
                float(await convert_amount(db, payment.amount, payment.currency, currency)),
                payment.note or "-",
            ]
            for payment, worker in payments
        ]

        audits = (
            await db.execute(
                select(AuditLog)
                .where(
                    AuditLog.group_id == group_id,
                    AuditLog.created_at >= start_date,
                    AuditLog.created_at <= end_date,
                )
                .order_by(desc(AuditLog.created_at))
                .limit(500)
            )
        ).scalars().all()
        actor_ids = {item.actor_user_id for item in audits if item.actor_user_id}
        actor_map = {}
        if actor_ids:
            actors = (await db.execute(select(User).where(User.id.in_(list(actor_ids))))).scalars().all()
            actor_map = {actor.id: _display_name(actor) for actor in actors}
        admin_payload["audit"] = [
            [item.created_at.strftime("%Y-%m-%d %H:%M"), actor_map.get(item.actor_user_id, "-"), item.action, item.entity_type, item.entity_id, item.payload or "-"]
            for item in audits
        ]

    return {
        "group_name": group.name,
        "period_label": _period_label(period_type, start_date, now, end_date),
        "currency": currency,
        "summary": summary_rows,
        "top_categories": top_categories,
        "transactions": transaction_rows,
        "debts": debt_rows,
        "debt_repayments": repayment_rows,
        "admin_mode": admin_mode,
        **admin_payload,
    }


def build_excel_workbook(payload: dict[str, Any]) -> bytes:
    wb = Workbook()
    styles = _styles()

    summary_ws = wb.active
    summary_ws.title = "Summary"
    _write_sheet_title(summary_ws, "Financial Monitoring Report", 3, styles)
    _write_table(summary_ws, start_row=3, headers=["Metric", "Value"], rows=payload["summary"], widths=[30, 28], numeric_cols={2})
    top_start = 5 + len(payload["summary"])
    summary_ws.cell(row=top_start, column=1, value="Top Categories").font = styles["section_font"]
    _write_table(
        summary_ws,
        start_row=top_start + 1,
        headers=["Category", "Amount", "%"],
        rows=payload["top_categories"] or [["-", 0, 0]],
        widths=[30, 18, 10],
        numeric_cols={2, 3},
    )

    tx_ws = wb.create_sheet("Transactions")
    _write_sheet_title(tx_ws, "Transactions", 9, styles)
    tx_rows = [
        [row["date"], row["type"], row["category"], row["funding"], row["source"], row["amount"], row["currency"], row["converted"], row["description"]]
        for row in payload["transactions"]
    ] or [["-", "-", "-", "-", "-", 0, payload["currency"], 0, "-"]]
    _write_table(
        tx_ws,
        start_row=3,
        headers=["Date", "Type", "Category", "Funding", "Debt Source", "Amount", "Currency", "Report Amount", "Description"],
        rows=tx_rows,
        widths=[20, 18, 24, 16, 22, 14, 10, 18, 40],
        numeric_cols={6, 8},
    )

    debts_ws = wb.create_sheet("Debts")
    _write_sheet_title(debts_ws, "Debts", 12, styles)
    debt_rows = payload["debts"] or [["-", "-", "-", "-", 0, payload["currency"], 0, 0, 0, "-", "-", "-"]]
    debt_rows = _append_total_row(debt_rows, "Total", {5, 7, 8, 9})
    _write_table(
        debts_ws,
        start_row=3,
        headers=["Created", "Kind", "Source", "Contact", "Amount", "Currency", "Remaining", "Used", "Available", "Status", "Description", "Paid At"],
        rows=debt_rows,
        widths=[20, 16, 24, 18, 14, 10, 14, 14, 14, 16, 34, 20],
        numeric_cols={5, 7, 8, 9},
    )

    repayments_ws = wb.create_sheet("Debt Repayments")
    _write_sheet_title(repayments_ws, "Debt Repayments", 7, styles)
    repayment_rows = payload["debt_repayments"] or [["-", "-", "-", 0, payload["currency"], 0, "-"]]
    repayment_rows = _append_total_row(repayment_rows, "Total", {4, 6})
    _write_table(
        repayments_ws,
        start_row=3,
        headers=["Date", "Debt", "Source", "Amount", "Currency", "Report Amount", "Note"],
        rows=repayment_rows,
        widths=[20, 28, 24, 14, 10, 18, 36],
        numeric_cols={4, 6},
    )

    if payload["admin_mode"]:
        sheets = [
            ("Users", ["Name", "Username", "Role", "Active", "Language", "Currency", "Joined At"], payload["users"], [24, 20, 14, 10, 12, 10, 20], set()),
            ("Transfers", ["Date", "Sender", "Recipient", "Amount", "Currency", "Remaining", "Spent", "Status", "Description"], payload["transfers"], [20, 24, 24, 14, 10, 14, 14, 14, 36], {4, 6, 7}),
            ("Workers", ["Name", "Role", "Payment Type", "Rate", "Currency", "Active", "Start Date", "Quantity", "Accrued", "Advance", "Paid", "Payable", "Status", "Phone", "Notes"], payload["workers"], [22, 18, 16, 14, 10, 10, 14, 12, 14, 14, 14, 14, 14, 16, 28], {4, 8, 9, 10, 11, 12}),
            ("Attendance", ["Date", "Worker", "Payment Type", "Status", "Units", "Comment"], payload["attendance"], [14, 24, 16, 16, 12, 34], {5}),
            ("Advances", ["Date", "Worker", "Amount", "Currency", "Report Amount", "Note"], payload["advances"], [14, 24, 14, 10, 18, 36], {3, 5}),
            ("Worker Payments", ["Date", "Worker", "Amount", "Currency", "Report Amount", "Note"], payload["worker_payments"], [14, 24, 14, 10, 18, 36], {3, 5}),
            ("Audit Log", ["Date", "Actor", "Action", "Entity", "Entity ID", "Payload"], payload["audit"], [20, 22, 28, 18, 20, 56], set()),
        ]
        for title, headers, rows, widths, numeric_cols in sheets:
            ws = wb.create_sheet(title)
            _write_sheet_title(ws, title, len(headers), styles)
            fallback = [["-"] * len(headers)]
            table_rows = rows or fallback
            if rows and numeric_cols:
                table_rows = _append_total_row(table_rows, "Total", numeric_cols)
            _write_table(ws, start_row=3, headers=headers, rows=table_rows, widths=widths, numeric_cols=numeric_cols)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream.read()


async def generate_excel_report(
    db: AsyncSession,
    user: User,
    period: str,
    *,
    date_from: date | None = None,
    date_to: date | None = None,
    filename_prefix: str = "report",
    include_admin_sheets: bool | None = None,
    group_id_override: int | None = None,
) -> tuple[bytes, str]:
    if include_admin_sheets is None:
        payload = await collect_excel_report_payload(db, user, period, date_from=date_from, date_to=date_to)
    else:
        group_id = int(group_id_override or await get_active_group_id(db, user))
        group = (await db.execute(select(Group).where(Group.id == group_id))).scalar_one()
        start_date, end_date, period_type = _period_bounds(period, date_from=date_from, date_to=date_to)
        payload = await collect_excel_report_payload_for_group(
            db,
            user,
            period,
            group_id=group_id,
            include_admin_sheets=include_admin_sheets,
            group=group,
            start_date=start_date,
            end_date=end_date,
            period_type=period_type,
            currency=normalize_currency(user.default_currency, "UZS"),
        )
    content = build_excel_workbook(payload)
    period_suffix = f"custom_{date_from:%Y%m%d}_{date_to:%Y%m%d}" if period == "custom" and date_from and date_to else period
    filename = f"{filename_prefix}_{_safe_filename_part(payload['group_name'])}_{period_suffix}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    return content, filename


async def generate_report_download(
    db: AsyncSession,
    user: User,
    period: str,
    *,
    date_from: date | None = None,
    date_to: date | None = None,
    filename_prefix: str = "report",
) -> tuple[bytes, str, str]:
    group_id = await get_active_group_id(db, user)
    admin_mode = await is_group_admin(db, user, group_id)
    if not admin_mode:
        content, filename = await generate_excel_report(
            db,
            user,
            period,
            date_from=date_from,
            date_to=date_to,
            filename_prefix=filename_prefix,
        )
        return content, filename, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    overview_content, overview_filename = await generate_excel_report(
        db,
        user,
        period,
        date_from=date_from,
        date_to=date_to,
        filename_prefix=f"{filename_prefix}_group",
        include_admin_sheets=True,
    )

    members = (
        await db.execute(
            select(User)
            .join(UserGroup, UserGroup.user_id == User.id)
            .where(UserGroup.group_id == group_id)
            .order_by(User.first_name.asc(), User.id.asc())
        )
    ).scalars().all()

    archive_stream = BytesIO()
    with zipfile.ZipFile(archive_stream, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(overview_filename, overview_content)
        manifest_lines = [f"Group report: {overview_filename}", ""]
        for member in members:
            member_name = _safe_filename_part(_display_name(member))
            member_content, member_filename = await generate_excel_report(
                db,
                member,
                period,
                date_from=date_from,
                date_to=date_to,
                filename_prefix=f"{filename_prefix}_{member_name}_{member.id}",
                include_admin_sheets=False,
                group_id_override=group_id,
            )
            archive.writestr(member_filename, member_content)
            manifest_lines.append(f"{member.id}: {member_filename}")
        archive.writestr("README.txt", "\n".join(manifest_lines))

    archive_stream.seek(0)
    group = await get_active_group(db, user)
    period_suffix = f"custom_{date_from:%Y%m%d}_{date_to:%Y%m%d}" if period == "custom" and date_from and date_to else period
    filename = f"{filename_prefix}_{_safe_filename_part(group.name)}_{period_suffix}_{datetime.now():%Y%m%d_%H%M%S}.zip"
    return archive_stream.read(), filename, "application/zip"
